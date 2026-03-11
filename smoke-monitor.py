import serial
import time
from twilio.rest import Client
from openpyxl import Workbook, load_workbook


# Twilio credentials (replace with your own)
TWILIO_ACCOUNT_SID = "YOUR_ACCOUNT_SID"
TWILIO_AUTH_TOKEN = "YOUR_AUTH_TOKEN"
TWILIO_PHONE_NUMBER = "+1234567890"
TARGET_PHONE_NUMBER = "+1234567890"


arduino_port = "COM4"
baud_rate = 9600


excel_file = r"D:\workvv\arduino\SmokeAlertLog.xlsx"


client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)


def initialize_excel():
    """Create and initialize the Excel file if it does not exist."""
    try:
        workbook = load_workbook(excel_file)
        print(f"Loaded existing file: {excel_file}")
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Alert Logs"
        sheet.append(
            [
                "Timestamp",
                "Date",
                "Time",
                "Smoke Level",
                "Alert Type",
                "Phone Number",
                "Location",
                "Device ID",
            ]
        )
        workbook.save(excel_file)
        print(f"Created new file: {excel_file}")
    return workbook


def log_to_excel(
    timestamp,
    date,
    time,
    smoke_level,
    alert_type,
    phone_number,
    location="Erode",
    device_id="Device_1",
):
    """Log alert details to Excel file."""
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    sheet.append(
        [
            timestamp,
            date,
            time,
            smoke_level,
            alert_type,
            phone_number,
            location,
            device_id,
        ]
    )
    workbook.save(excel_file)
    print(
        f"Logged alert to Excel: {timestamp}, {smoke_level}, {alert_type}, {phone_number}"
    )


def send_sms():
    """Send an SMS alert using Twilio."""
    try:
        message = client.messages.create(
            body="ALERT: Smoke detected! Please evacuate immediately.",
            from_=TWILIO_PHONE_NUMBER,
            to=TARGET_PHONE_NUMBER,
        )
        print("SMS alert sent.")
        return "SMS"
    except Exception as e:
        print(f"Failed to send SMS: {e}")
        return None


def make_call():
    """Make a call alert using Twilio."""
    try:
        call = client.calls.create(
            to=TARGET_PHONE_NUMBER,
            from_=TWILIO_PHONE_NUMBER,
            url="http://demo.twilio.com/docs/voice.xml",
        )
        print(f"Call alert initiated. Call SID: {call.sid}")
        return "Call"
    except Exception as e:
        print(f"Failed to make a call: {e}")
        return None


print("Initializing serial connection...")
try:
    arduino = serial.Serial(arduino_port, baud_rate, timeout=1)
    time.sleep(2)  # Allow time for Arduino to stabilize
    print("Serial connection established!")
except serial.SerialException as e:
    print(f"Error: Could not open port {arduino_port}. {e}")
    exit(1)

# Initialize Excel file
initialize_excel()

# Main Loop to Listen for Smoke Level Data
print("Listening for commands from Arduino...")
try:
    while True:
        if arduino.in_waiting > 0:
            # Read and process the smoke level
            message = arduino.readline().decode().strip()

            if message.isdigit():  # Check if the message is numeric
                smoke_level = int(message)
                print(f"Smoke Level: {smoke_level}")

                # Trigger alerts if smoke level exceeds threshold
                if smoke_level > 300:  # Adjust threshold as needed
                    print("Smoke detected! Triggering alerts...")

                    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
                    date = time.strftime("%Y-%m-%d")
                    current_time = time.strftime("%H:%M:%S")

                    # Send SMS and log to Excel
                    alert_type_sms = send_sms()
                    if alert_type_sms:
                        log_to_excel(
                            timestamp,
                            date,
                            current_time,
                            smoke_level,
                            alert_type_sms,
                            TARGET_PHONE_NUMBER,
                        )

                    # Make call and log to Excel
                    alert_type_call = make_call()
                    if alert_type_call:
                        log_to_excel(
                            timestamp,
                            date,
                            current_time,
                            smoke_level,
                            alert_type_call,
                            TARGET_PHONE_NUMBER,
                        )

                    print("Alert process completed.")  # Confirm after both alerts
                    time.sleep(10)  # Delay to prevent duplicate alerts

        time.sleep(1)  # Polling delay

except KeyboardInterrupt:
    print("Program terminated by user.")

finally:
    arduino.close()
    print("Serial connection closed.")
